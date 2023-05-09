import functools
import string
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook


def data_merger(data_columnas, *data_to_merge):
    '''
    Hace un merge_update de todos los ratos reunidados en la tabla
    '''
    resultado = []
    
    flat_list = [data for sb_data in data_to_merge for data in sb_data] + data_columnas
    
    data_columna = [col['columna'] for col in data_columnas]

    for col in data_columna:
        res = None

        item = [data for data in flat_list if data['columna'] == col]

        if item != [] : 
            res = functools.reduce(lambda a,b : {**a,**b}, item)
            resultado.append(res)
        
    return resultado


def gather_data(completitud ,exactitud , data_columnas ):

    gather_exactitud = []
    gather_completitud = []

    if completitud is not None:

        data = completitud[['COLUMNA','REGISTROS TOTALES','CANTIDAD DE INCOMPLETOS']].to_numpy()
        

        for val in data:

            
            cols = {}
            cols['columna'] = val[0] 
            cols['total_registros']  = val[1]
            cols['total_nulos']  =  val[2]

            gather_completitud.append(cols)
    else:
        print('No hay data de completitud')
        return


    if exactitud is not None:
        data = exactitud[['COLUMNA','INEXACTO']].to_numpy()


        for val in data:
            
            data = {}
            data['columna'] = val[0]
            data['total_errores']  = val[1]
            
            gather_exactitud.append(data)
    else:
        print('No hay data de exactitud')
        gather_exactitud = [{'columna': col['columna'] , 'total_errores':None} for col in data_columnas]
       
    
    resultado_mix = data_merger(data_columnas, gather_exactitud , gather_completitud )
    
    return resultado_mix


def write_data(index , data , database ,ws):
    '''
    Columna B NOMBRE CAMPO
    Columna C TIPO CAMPO 
    Columna D BASE DATOS
    Columna E TOTAL REGISTROS 
    Columna F TOTAL NULOS
    Columna G TOTAL ERRORES
    '''


    ws[f'B{index}'].value = data.get('columna',None)
    ws[f'C{index}'].value = data.get('type',None)
    ws[f'D{index}'].value = database
    ws[f'E{index}'].value = data['total_registros']
    ws[f'F{index}'].value = data.get('total_nulos' ,None)
    ws[f'G{index}'].value = data.get('total_errores',None)


    return None
    


def insercion_data(insertion_data):

## TODO Cambiar la forma de como acepta los datos y agregar a la inserción multiple

    workbook = load_workbook(insertion_data.get('template_wb_matriz'))
    output = insertion_data.get('output')
    worksheet = workbook.active

    starting_index = 6

    
    value = insertion_data.get('data_builder',None)
    database = value.get('database',None) if value is not None else None
    completitud = insertion_data.get('completitud')
    exactitud = insertion_data.get('exactitud')
    data_columnas = insertion_data.get('data_columnas')

    list_data = gather_data(completitud, exactitud ,data_columnas)
    
    for index,data in enumerate(list_data,start=starting_index):
        write_data(index, data, database , worksheet)
    
    starting_index = index + 2
    
    
    try:
            
        workbook.save(output)

    except Exception as e:
        print(e)
    finally:
        print(f"Se realizo efectivamente la carga de datos en la carpeta output con el nombre:\n{output}")

    return 





